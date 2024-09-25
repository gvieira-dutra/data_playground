# import os
# from openpyxl import load_workbook
# from selenium import webdriver
# from selenium.webdriver.common.by import By
# import time
# from selenium import webdriver
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC

# def main():
#     # Define the path to the Excel file
#     file_path = r'C:\R\History\pgx.xlsx'

#     # Create a list to hold the links from the History_link column
#     history_urls = []

#     # Check if file exists
#     if not os.path.exists(file_path):
#         print("The file was not found.")
#         return

#     # Load the Excel file
#     workbook = load_workbook(file_path)
    
#     # Get the worksheet named "All_with_2024_updates"
#     if "All_with_2024_updates" not in workbook.sheetnames:
#         print("The worksheet 'All_with_2024_updates' was not found.")
#         return
    
#     worksheet = workbook["All_with_2024_updates"]

#     # Find the "History_link" column by its header name
#     history_link_column = None

#     # Loop through the first row to find the column index for "History_link"
#     for cell in worksheet[1]:
#         if cell.value and cell.value.lower() == "history_link":
#             history_link_column = cell.column
#             break

#     if history_link_column is None:
#         print("The 'History_link' column was not found.")
#         return

#     # Loop through all rows starting from the second row
#     for row in worksheet.iter_rows(min_row=2):
#         cell = row[history_link_column - 1]  # Adjust for zero-based index
#         link = cell.value

#         if link and isinstance(link, str) and link.strip():
#             history_urls.append(link)

#     # Output the links (optional)
#     #for i, link in enumerate(history_urls, start=1):
#         #print(f"{i} {link}")

#     # Set up the WebDriver (make sure you have downloaded ChromeDriver)
#     driver = webdriver.Chrome()


#     # Dictionary to store history updates for each link
#     history_data = {}

#     for url in history_urls:
#         driver.get(url)

#        # Wait for the History header to load
#         history_header = WebDriverWait(driver, 10).until(
#             EC.presence_of_element_located((By.XPATH, "//h3[contains(text(), 'History')]"))
#         )

#         try: 
#             # Find the table after the History header
#             table = history_header.find_element(By.XPATH, "./following-sibling::div//table")  # Adjust based on table structure
        
#             # Retrieve all rows in the table
#             rows = table.find_elements(By.TAG_NAME, "tr")

#                  # Extract the relevant data from rows for 2024
#             history_updates = []
#             for row in rows:
#                 cells = row.find_elements(By.TAG_NAME, "td")
#                 if len(cells) > 0:
#                     date = cells[0].text
#                     if "2024" in date:
#                         history_updates.append({
#                             "date": date,
#                             "type": cells[1].text,
#                             "comment": cells[2].text
#                         })
        
#             # Store history updates in the dictionary
#             if history_updates:
#                 history_data[url] = history_updates
    
#         except Exception as e:
#             print(f"Error retrieving data from {url}: {str(e)}")

#         # Update the "Notes_Sept2024" column based on history_data
#     for url, updates in history_data.items():
#         # Find the row corresponding to the current URL
#         for row in worksheet.iter_rows(min_row=2):
#             cell = row[history_link_column - 1]  # Adjust for zero-based index
#             if cell.value == url:
#                 # Join the updates into a string (customize this format as needed)
#                 notes = "\n".join([f"{update['date']}: {update['type']} - {update['comment']}" for update in updates])
#                 # Update the Notes_Sept2024 column
#                 row[notes_sept_column - 1].value = notes
#                 break  # Exit the loop after updating the corresponding row

#     # Save the changes to the Excel file
#     workbook.save(file_path)

#     # Close the browser
#     driver.quit()










# if __name__ == "__main__":
#     main()


import os
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
import time
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

def main():
    # Define the path to the Excel file
    file_path = r'C:\R\History\pgx.xlsx'

    # Create a list to hold the links from the History_link column
    history_urls = []

    # Check if file exists
    if not os.path.exists(file_path):
        print("The file was not found.")
        return

    # Load the Excel file
    workbook = load_workbook(file_path)
    
    # Get the worksheet named "All_with_2024_updates"
    if "All_with_2024_updates" not in workbook.sheetnames:
        print("The worksheet 'All_with_2024_updates' was not found.")
        return
    
    worksheet = workbook["All_with_2024_updates"]

    # Find the "History_link" and "Notes_Sept2024" columns by their header names
    history_link_column = None
    notes_sept_column = None

    # Loop through the first row to find the column indices for "History_link" and "Notes_Sept2024"
    for cell in worksheet[1]:
        if cell.value and cell.value.lower() == "history_link":
            history_link_column = cell.column
        elif cell.value and cell.value.lower() == "notes_sept2024":
            notes_sept_column = cell.column

    if history_link_column is None:
        print("The 'History_link' column was not found.")
        return
    if notes_sept_column is None:
        print("The 'Notes_Sept2024' column was not found.")
        return

    # Loop through all rows starting from the second row
    for row in worksheet.iter_rows(min_row=2):
        cell = row[history_link_column - 1]  # Adjust for zero-based index
        link = cell.value

        if link and isinstance(link, str) and link.strip():
            history_urls.append(link)

    # Set up the WebDriver (make sure you have downloaded ChromeDriver)
    driver = webdriver.Chrome()

    # Dictionary to store history updates for each link
    history_data = {}

    for url in history_urls:
        driver.get(url)

        # Wait for the History header to load
        history_header = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//h3[contains(text(), 'History')]"))
        )

        try:
            # Find the table after the History header
            table = history_header.find_element(By.XPATH, "./following-sibling::div//table")  # Adjust based on table structure
        
            # Retrieve all rows in the table
            rows = table.find_elements(By.TAG_NAME, "tr")

            # Extract the relevant data from rows for 2024
            history_updates = []
            for row in rows:
                cells = row.find_elements(By.TAG_NAME, "td")
                if len(cells) > 0:
                    date = cells[0].text
                    if "2024" in date:
                        history_updates.append({
                            "date": date,
                            "type": cells[1].text,
                            "comment": cells[2].text
                        })
        
            # Store history updates in the dictionary
            if history_updates:
                history_data[url] = history_updates
    
        except Exception as e:
            print(f"Error retrieving data from {url}: {str(e)}")

    # Update the "Notes_Sept2024" column based on history_data
    for url, updates in history_data.items():
        # Find the row corresponding to the current URL
        for row in worksheet.iter_rows(min_row=2):
            cell = row[history_link_column - 1]  # Adjust for zero-based index
            if cell.value == url:
                # Join the updates into a string (customize this format as needed)
                notes = "\n".join([f"{update['date']}: {update['type']} - {update['comment']}" for update in updates])
                # Update the Notes_Sept2024 column
                row[notes_sept_column - 1].value = notes
                break  # Exit the loop after updating the corresponding row

    # Save the changes to the Excel file
    workbook.save(file_path)

    # Close the browser
    driver.quit()

if __name__ == "__main__":
    main()
